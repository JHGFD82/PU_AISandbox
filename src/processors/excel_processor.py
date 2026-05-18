"""Excel file processor: reads .xlsx/.xls files and converts sheets to text pages."""

import logging
from typing import List

from .base_text_processor import BaseTextProcessor
from .constants import DEFAULT_PAGE_SIZE


class ExcelProcessor(BaseTextProcessor):
    """Handles extraction of text from Excel workbooks (.xlsx / .xls)."""

    @staticmethod
    def process_excel_with_pages(file_path: str, target_page_size: int = DEFAULT_PAGE_SIZE) -> List[str]:
        """Read an Excel workbook and return its content as logical text pages.

        Each sheet is rendered as a header line followed by tab-separated rows.
        Multiple sheets are separated by blank lines.  The resulting text is then
        split into logical pages the same way as :meth:`TxtProcessor.process_txt_with_pages`.

        Args:
            file_path: Absolute path to the .xlsx or .xls file.
            target_page_size: Target number of characters per logical page.

        Returns:
            List of strings, each representing a logical page of content.
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to process Excel files. "
                "Install it with: pip install openpyxl"
            ) from exc

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise Exception(f"Failed to open Excel file '{file_path}': {exc}") from exc

        processor = ExcelProcessor()
        sheet_blocks: List[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in cells):
                    rows.append("\t".join(cells))

            if rows:
                block = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
                sheet_blocks.append(block)

        wb.close()

        if not sheet_blocks:
            logging.warning("No data found in Excel file '%s'", file_path)
            return [""]

        full_text = "\n\n".join(sheet_blocks)
        paragraphs = processor.parse_text_into_paragraphs(full_text)
        pages = processor.split_text_into_pages(paragraphs, target_page_size)

        logging.info("Split Excel file into %d logical page(s)", len(pages))
        return pages

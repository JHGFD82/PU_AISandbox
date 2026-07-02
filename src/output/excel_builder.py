"""Excel output builder: writes AI response content to .xlsx files."""

import logging
from pathlib import Path
from typing import List, Optional

from ._output_utils import _emit_message, _parse_md_table_block, save_to_text_file


def save_to_excel(
    content: str,
    output_path: str,
    label: str = "Output",
) -> None:
    """Save the AI's response text to an Excel workbook (.xlsx file).

    Any tables in the response are placed on their own worksheet ("Table 1",
    "Table 2", and so on) so the rows and columns display properly in Excel,
    with the header row shown in bold. Any surrounding prose (text that isn't
    part of a table) is written to its own "Text" worksheet, one paragraph
    per row. If the response contains no tables at all, every paragraph is
    written to a single "Content" worksheet instead.

    This function relies on the ``openpyxl`` package. If it isn't installed,
    the response is saved as a plain .txt file instead, with a note printed
    explaining why.

    Args:
        content: The AI's response text to save.
        output_path: The file path to write to, e.g. ``'response.xlsx'``.
        label: A short description used in the saved-confirmation message
               shown to the user (e.g. ``'Translation'`` or ``'Response'``).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        _emit_message(
            f"Warning: openpyxl not installed — saving as .txt instead of .xlsx",
            level=logging.WARNING,
        )
        text_path = str(Path(output_path).with_suffix(".txt"))
        save_to_text_file(content, text_path, label)
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Split content into prose blocks and Markdown table blocks ──────────
    import re
    table_pattern = re.compile(
        r'((?:^\|.+\n)+)',
        re.MULTILINE,
    )

    table_index = 0
    prose_parts: List[str] = []
    last_end = 0

    for match in table_pattern.finditer(content):
        prose_before = content[last_end:match.start()].strip()
        if prose_before:
            prose_parts.append(prose_before)

        table_block = match.group(1).strip()
        grid = _parse_md_table_block(table_block)
        if grid is not None:
            table_index += 1
            ws = wb.create_sheet(title=f"Table {table_index}")
            for row_idx, row in enumerate(grid, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    if row_idx == 1:
                        cell.font = Font(bold=True)
        else:
            # Not a recognized table — treat as prose
            prose_parts.append(table_block)

        last_end = match.end()

    trailing_prose = content[last_end:].strip()
    if trailing_prose:
        prose_parts.append(trailing_prose)

    # ── Write prose ────────────────────────────────────────────────────────
    prose_text = "\n\n".join(prose_parts).strip()
    if prose_text:
        ws_text = wb.create_sheet(title="Text")
        for row_idx, para in enumerate(
            (p.strip() for p in prose_text.split("\n") if p.strip()), start=1
        ):
            ws_text.cell(row=row_idx, column=1, value=para)

    # ── Fallback: no tables and no prose captured separately ───────────────
    if not wb.sheetnames:
        ws_content = wb.create_sheet(title="Content")
        for row_idx, para in enumerate(
            (p.strip() for p in content.split("\n") if p.strip()), start=1
        ):
            ws_content.cell(row=row_idx, column=1, value=para)

    try:
        wb.save(output_path)
        _emit_message(f"{label} saved to {output_path}")
    except Exception as exc:
        _emit_message(
            f"Error saving Excel file: {exc} — falling back to .txt",
            level=logging.ERROR,
        )
        text_path = str(Path(output_path).with_suffix(".txt"))
        save_to_text_file(content, text_path, label)

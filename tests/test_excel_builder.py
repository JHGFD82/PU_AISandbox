"""Tests for ExcelBuilder (save_to_excel)."""

from pathlib import Path

import pytest

from src.output.excel_builder import save_to_excel


openpyxl = pytest.importorskip("openpyxl")


class TestSaveToExcel:

    def test_creates_file(self, tmp_path):
        out = str(tmp_path / "out.xlsx")
        save_to_excel("Hello world", out, label="Test")
        assert Path(out).exists()

    def test_plain_text_goes_to_content_or_text_sheet(self, tmp_path):
        out = str(tmp_path / "plain.xlsx")
        save_to_excel("First line\nSecond line\nThird line", out, label="Test")
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames  # at least one sheet

    def test_markdown_table_creates_table_sheet(self, tmp_path):
        content = "| Name | Age |\n| --- | --- |\n| Alice | 30 |\n| Bob | 25 |"
        out = str(tmp_path / "table.xlsx")
        save_to_excel(content, out, label="Test")
        wb = openpyxl.load_workbook(out)
        assert any("Table" in name for name in wb.sheetnames)

    def test_markdown_table_header_is_bold(self, tmp_path):
        content = "| Name | Age |\n| --- | --- |\n| Alice | 30 |"
        out = str(tmp_path / "bold.xlsx")
        save_to_excel(content, out, label="Test")
        wb = openpyxl.load_workbook(out)
        table_sheet = next(ws for ws in wb.worksheets if "Table" in ws.title)
        header_cell = table_sheet.cell(row=1, column=1)
        assert header_cell.font.bold

    def test_prose_and_table_together(self, tmp_path):
        content = (
            "Introduction paragraph.\n\n"
            "| A | B |\n| --- | --- |\n| 1 | 2 |\n\n"
            "Conclusion paragraph."
        )
        out = str(tmp_path / "mixed.xlsx")
        save_to_excel(content, out, label="Test")
        wb = openpyxl.load_workbook(out)
        names = wb.sheetnames
        assert any("Table" in n for n in names)
        assert any(n in ("Text", "Content") for n in names)

    def test_fallback_to_txt_when_openpyxl_missing(self, tmp_path):
        import builtins
        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return real_import(name, *args, **kwargs)

        out = str(tmp_path / "fallback.xlsx")
        with pytest.MonkeyPatch.context() as mp:
            mp.setattr("builtins.__import__", fake_import)
            save_to_excel("Some content", out, label="Test")
        # Fallback writes a .txt file
        assert Path(str(tmp_path / "fallback.txt")).exists()

    def test_pipe_block_without_separator_treated_as_prose(self, tmp_path):
        # A block that matches the pipe regex but lacks a separator row (---).
        # _parse_md_table_block returns None → treated as prose (line 69).
        content = "| foo | bar |\n| baz | qux |"
        out = str(tmp_path / "nontable.xlsx")
        save_to_excel(content, out, label="Test")
        wb = openpyxl.load_workbook(out)
        # Should produce a Text sheet (prose) not a Table sheet
        assert not any("Table" in name for name in wb.sheetnames)

    def test_empty_content_creates_content_sheet(self, tmp_path):
        # Empty string produces no prose and no tables, so the fallback
        # "Content" sheet is created (lines 88-92).
        out = str(tmp_path / "empty.xlsx")
        save_to_excel("", out, label="Test")
        wb = openpyxl.load_workbook(out)
        assert "Content" in wb.sheetnames

    def test_save_error_falls_back_to_txt(self, tmp_path):
        # Force wb.save() to raise so the except branch (lines 97-103) runs.
        out = str(tmp_path / "err.xlsx")
        with pytest.MonkeyPatch.context() as mp:
            import openpyxl as _opx

            def bad_save(self, filename):
                raise OSError("disk full")

            mp.setattr(_opx.Workbook, "save", bad_save)
            save_to_excel("Some content", out, label="Test")
        assert Path(str(tmp_path / "err.txt")).exists()
